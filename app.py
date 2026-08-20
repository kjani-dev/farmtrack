from flask import Flask, render_template, request, redirect, url_for, Response, session
import json
import os
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import io

app = Flask(__name__)
app.secret_key = 'farmtrack-dev-secret-change-later'

USERS_FILE = '/home/farmtrackdev/farmtrack/data/users.json'
DATA_DIR = '/home/farmtrackdev/farmtrack/data'

def load_users():
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, 'r') as f:
            return json.load(f)
    return {}

def save_users(users):
    with open(USERS_FILE, 'w') as f:
        json.dump(users, f)

def user_data_file(username):
    return f'{DATA_DIR}/entries_{username}.json'

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'username' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        username = request.form['username'].strip().lower()
        password = request.form['password']
        users = load_users()
        if username in users:
            return render_template('signup.html', error='That username is already taken.')
        users[username] = generate_password_hash(password)
        save_users(users)
        session['username'] = username
        return redirect(url_for('dashboard'))
    return render_template('signup.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username'].strip().lower()
        password = request.form['password']
        users = load_users()
        if username not in users or not check_password_hash(users[username], password):
            return render_template('login.html', error='Incorrect username or password.')
        session['username'] = username
        return redirect(url_for('dashboard'))
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('username', None)
    return redirect(url_for('login'))

@app.route('/')
@login_required
def dashboard():
    return render_template('dashboard.html')

@app.route('/log', methods=['GET', 'POST'])
@login_required
def log_cost():
    data_file = user_data_file(session['username'])
    if request.method == 'POST':
        field_name = request.form['field_name']
        if field_name == '__new__':
            field_name = request.form['new_field_name'].strip()
        input_type = request.form['input_type']
        if input_type == '__custom__':
            input_type = request.form.get('custom_input_type', '').strip()
        entry = {
            'field_name': field_name,
            'acres': request.form['acres'],
            'crop_type': request.form.get('crop_type', ''),
            'input_type': input_type,
            'cost': request.form['cost'],
            'quantity': request.form.get('quantity', ''),
            'quantity_unit': request.form.get('quantity_unit', ''),
            'herbicide_group': request.form.get('herbicide_group', ''),
            'date': request.form['date']
        }
        if os.path.exists(data_file):
            with open(data_file, 'r') as f:
                entries = json.load(f)
        else:
            entries = []
        entries.append(entry)
        with open(data_file, 'w') as f:
            json.dump(entries, f)
        return redirect(url_for('dashboard'))
    if os.path.exists(data_file):
        with open(data_file, 'r') as f:
            entries = json.load(f)
        existing_fields = sorted(set(
            (e['field_name'], e['acres']) for e in entries
        ), key=lambda x: x[0])
    else:
        existing_fields = []
    return render_template('log_cost.html', existing_fields=existing_fields)

@app.route('/view')
@login_required
def view_costs():
    data_file = user_data_file(session['username'])
    if os.path.exists(data_file):
        with open(data_file, 'r') as f:
            entries = json.load(f)
    else:
        entries = []
    fields = {}
    for real_index, entry in enumerate(entries):
        name = entry['field_name']
        if name not in fields:
            fields[name] = {'acres': entry['acres'], 'crop_type': '', 'costs': []}
        if entry.get('crop_type'):
            fields[name]['crop_type'] = entry['crop_type']
        fields[name]['costs'].append({
            'input_type': entry['input_type'],
            'cost': float(entry['cost']),
            'date': entry['date'],
            'quantity': entry.get('quantity', ''),
            'herbicide_group': entry.get('herbicide_group', ''),
            'real_index': real_index
        })
    return render_template('view_costs.html', fields=fields)

@app.route('/summary')
@login_required
def summary():
    data_file = user_data_file(session['username'])
    if os.path.exists(data_file):
        with open(data_file, 'r') as f:
            entries = json.load(f)
    else:
        entries = []
    fields = {}
    for entry in entries:
        name = entry['field_name']
        if name not in fields:
            fields[name] = {'acres': float(entry['acres']), 'crop_type': '', 'total_cost': 0}
        if entry.get('crop_type'):
            fields[name]['crop_type'] = entry['crop_type']
        fields[name]['total_cost'] += float(entry['cost'])
    for field in fields.values():
        field['cost_per_acre'] = field['total_cost'] / field['acres']
    input_total = sum(f['total_cost'] for f in fields.values())

    fixed_file = f'{DATA_DIR}/fixedcosts_{session["username"]}.json'
    if os.path.exists(fixed_file):
        with open(fixed_file, 'r') as f:
            fixed_entries = json.load(f)
    else:
        fixed_entries = []
    fixed_total = sum(float(e['cost']) for e in fixed_entries)

    grand_total = input_total + fixed_total
    return render_template('summary.html', fields=fields, grand_total=grand_total, input_total=input_total, fixed_total=fixed_total)

@app.route('/delete/<int:index>')
@login_required
def delete_entry(index):
    data_file = user_data_file(session['username'])
    if os.path.exists(data_file):
        with open(data_file, 'r') as f:
            entries = json.load(f)
    else:
        entries = []
    if 0 <= index < len(entries):
        entries.pop(index)
    with open(data_file, 'w') as f:
        json.dump(entries, f)
    return redirect(url_for('view_costs'))

@app.route('/edit/<int:index>', methods=['GET', 'POST'])
@login_required
def edit_entry(index):
    data_file = user_data_file(session['username'])
    if os.path.exists(data_file):
        with open(data_file, 'r') as f:
            entries = json.load(f)
    else:
        entries = []
    if request.method == 'POST':
        input_type = request.form['input_type']
        if input_type == '__custom__':
            input_type = request.form.get('custom_input_type', '').strip()

        crop_type = request.form.get('crop_type', '')
        if crop_type == '__custom_crop__':
            crop_type = request.form.get('custom_crop_type', '').strip()

        herbicide_group = request.form.get('herbicide_group', '')
        if herbicide_group == '__custom_herb__':
            herbicide_group = request.form.get('custom_herbicide_group', '').strip()

        entries[index] = {
            'field_name': request.form['field_name'],
            'acres': request.form['acres'],
            'crop_type': crop_type,
            'input_type': input_type,
            'cost': request.form['cost'],
            'quantity': request.form.get('quantity', ''),
            'quantity_unit': request.form.get('quantity_unit', ''),
            'herbicide_group': herbicide_group,
            'date': request.form['date']
        }
        with open(data_file, 'w') as f:
            json.dump(entries, f)
        return redirect(url_for('view_costs'))
    entry = entries[index]
    known_types = ['Seed','46-0-0 (Urea — Nitrogen)','82-0-0 (NH3 — Nitrogen)','28-0-0 (Liquid — Nitrogen)','11-52-0 (MAP — Phosphate)','0-0-60 (Potash)','20.5-0-0-24 (Sulphur)','MES S15 (13-33-0-15)','Herbicide','Fungicide','Insecticide','PPE',"Labour (Arm's Length)",'Crop Insurance Premium','Fuel & Oil','Equipment Repairs','Custom Work']
    entry['is_custom_input'] = entry.get('input_type') not in known_types
    known_herb_groups = ['Group 1','Group 2','Group 4','Group 6','Group 9','Group 10','Group 14','Group 15','Group 27']
    herb_val = entry.get('herbicide_group', '')
    entry['is_custom_herbicide'] = bool(herb_val) and herb_val not in known_herb_groups

    known_crops = ['Wheat','Canola','Barley','Corn','Soybeans','Oats','Flax','Sunflowers']
    crop_val = entry.get('crop_type', '')
    entry['is_custom_crop'] = bool(crop_val) and crop_val not in known_crops

    return render_template('edit_entry.html', entry=entry, index=index)

@app.route('/equipment', methods=['GET', 'POST'])
@login_required
def equipment():
    data_file = f'{DATA_DIR}/equipment_{session["username"]}.json'
    if request.method == 'POST':
        entry = {
            'name': request.form['name'],
            'equipment_type': request.form.get('equipment_type', ''),
            'model': request.form.get('model', ''),
            'year': request.form.get('year', ''),
            'serial_number': request.form.get('serial_number', '')
        }
        if os.path.exists(data_file):
            with open(data_file, 'r') as f:
                equip_list = json.load(f)
        else:
            equip_list = []
        entry['id'] = len(equip_list)
        equip_list.append(entry)
        with open(data_file, 'w') as f:
            json.dump(equip_list, f)
        return redirect(url_for('equipment'))
    if os.path.exists(data_file):
        with open(data_file, 'r') as f:
            equip_list = json.load(f)
    else:
        equip_list = []
    return render_template('equipment.html', equip_list=equip_list)

@app.route('/equipment/<int:equip_id>')
@login_required
def equipment_detail(equip_id):
    equip_file = f'{DATA_DIR}/equipment_{session["username"]}.json'
    maint_file = f'{DATA_DIR}/maintenance_{session["username"]}.json'

    if os.path.exists(equip_file):
        with open(equip_file, 'r') as f:
            equip_list = json.load(f)
    else:
        equip_list = []
    equip = next((e for e in equip_list if e['id'] == equip_id), None)
    if not equip:
        return redirect(url_for('equipment'))

    if os.path.exists(maint_file):
        with open(maint_file, 'r') as f:
            all_maint = json.load(f)
    else:
        all_maint = []
    maint_records = [m for i, m in enumerate(all_maint) if m['equipment_id'] == equip_id]
    for i, m in enumerate(all_maint):
        m['real_index'] = i
    maint_records = [m for m in all_maint if m['equipment_id'] == equip_id]

    return render_template('equipment_detail.html', equip=equip, maint_records=maint_records)

@app.route('/equipment/<int:equip_id>/log', methods=['POST'])
@login_required
def log_maintenance(equip_id):
    maint_file = f'{DATA_DIR}/maintenance_{session["username"]}.json'
    if os.path.exists(maint_file):
        with open(maint_file, 'r') as f:
            all_maint = json.load(f)
    else:
        all_maint = []
    service_type = request.form['service_type']
    if service_type == '__custom__':
        service_type = request.form.get('custom_service_type', '').strip()
    entry = {
        'equipment_id': equip_id,
        'service_type': service_type,
        'hours': request.form.get('hours', ''),
        'date': request.form['date'],
        'notes': request.form.get('notes', '')
    }
    all_maint.append(entry)
    with open(maint_file, 'w') as f:
        json.dump(all_maint, f)
    return redirect(url_for('equipment_detail', equip_id=equip_id))

@app.route('/equipment/<int:equip_id>/delete-log/<int:index>')
@login_required
def delete_maintenance(equip_id, index):
    maint_file = f'{DATA_DIR}/maintenance_{session["username"]}.json'
    if os.path.exists(maint_file):
        with open(maint_file, 'r') as f:
            all_maint = json.load(f)
        if 0 <= index < len(all_maint):
            all_maint.pop(index)
        with open(maint_file, 'w') as f:
            json.dump(all_maint, f)
    return redirect(url_for('equipment_detail', equip_id=equip_id))

@app.route('/bins', methods=['GET', 'POST'])
@login_required
def bins():
    data_file = f'{DATA_DIR}/bins_{session["username"]}.json'
    if request.method == 'POST':
        entry = {
            'bin_name': request.form['bin_name'],
            'crop_type': request.form.get('crop_type', ''),
            'bushels': request.form['bushels'],
            'moisture': request.form.get('moisture', ''),
            'date': request.form['date']
        }
        if os.path.exists(data_file):
            with open(data_file, 'r') as f:
                entries = json.load(f)
        else:
            entries = []
        entries.append(entry)
        with open(data_file, 'w') as f:
            json.dump(entries, f)
        return redirect(url_for('bins'))
    if os.path.exists(data_file):
        with open(data_file, 'r') as f:
            entries = json.load(f)
    else:
        entries = []
    for i, e in enumerate(entries):
        e['real_index'] = i
        try:
            m = float(e.get('moisture', 0))
            e['moisture_safe'] = 14.5 <= m <= 15.5
        except (ValueError, TypeError):
            e['moisture_safe'] = None
    total_bushels = sum(float(e['bushels']) for e in entries)
    return render_template('bins.html', entries=entries, total_bushels=total_bushels)

@app.route('/bins/delete/<int:index>')
@login_required
def delete_bin(index):
    data_file = f'{DATA_DIR}/bins_{session["username"]}.json'
    if os.path.exists(data_file):
        with open(data_file, 'r') as f:
            entries = json.load(f)
    else:
        entries = []
    if 0 <= index < len(entries):
        entries.pop(index)
    with open(data_file, 'w') as f:
        json.dump(entries, f)
    return redirect(url_for('bins'))

@app.route('/rotation')
@login_required
def rotation():
    data_file = user_data_file(session['username'])
    if os.path.exists(data_file):
        with open(data_file, 'r') as f:
            entries = json.load(f)
    else:
        entries = []

    field_years = {}
    for entry in entries:
        field = entry['field_name']
        crop = entry.get('crop_type', '')
        date = entry.get('date', '')
        if not crop or not date:
            continue
        year = date[:4]
        if field not in field_years:
            field_years[field] = {}
        field_years[field][year] = crop

    fields_rotation = {}
    for field, years in field_years.items():
        sorted_years = sorted(years.keys(), reverse=True)
        fields_rotation[field] = [(y, years[y]) for y in sorted_years]

    return render_template('rotation.html', fields_rotation=fields_rotation)

@app.route('/fixed-costs', methods=['GET', 'POST'])
@login_required
def fixed_costs():
    data_file = f'{DATA_DIR}/fixedcosts_{session["username"]}.json'
    if request.method == 'POST':
        category = request.form['category']
        if category == '__custom__':
            category = request.form.get('custom_category', '').strip()
        entry = {
            'category': category,
            'cost': request.form['cost'],
            'notes': request.form.get('notes', ''),
            'date': request.form['date']
        }
        if os.path.exists(data_file):
            with open(data_file, 'r') as f:
                entries = json.load(f)
        else:
            entries = []
        entries.append(entry)
        with open(data_file, 'w') as f:
            json.dump(entries, f)
        return redirect(url_for('fixed_costs'))
    if os.path.exists(data_file):
        with open(data_file, 'r') as f:
            entries = json.load(f)
    else:
        entries = []
    for i, e in enumerate(entries):
        e['real_index'] = i
    total_fixed = sum(float(e['cost']) for e in entries)
    return render_template('fixed_costs.html', entries=entries, total_fixed=total_fixed)

@app.route('/fixed-costs/delete/<int:index>')
@login_required
def delete_fixed_cost(index):
    data_file = f'{DATA_DIR}/fixedcosts_{session["username"]}.json'
    if os.path.exists(data_file):
        with open(data_file, 'r') as f:
            entries = json.load(f)
    else:
        entries = []
    if 0 <= index < len(entries):
        entries.pop(index)
    with open(data_file, 'w') as f:
        json.dump(entries, f)
    return redirect(url_for('fixed_costs'))

@app.route('/download')
@login_required
def download_pdf():
    data_file = user_data_file(session['username'])
    if os.path.exists(data_file):
        with open(data_file, 'r') as f:
            entries = json.load(f)
    else:
        entries = []
    fields = {}
    for entry in entries:
        name = entry['field_name']
        if name not in fields:
            fields[name] = {'acres': float(entry['acres']), 'crop_type': '', 'total_cost': 0}
        if entry.get('crop_type'):
            fields[name]['crop_type'] = entry['crop_type']
        fields[name]['total_cost'] += float(entry['cost'])
    for field in fields.values():
        field['cost_per_acre'] = field['total_cost'] / field['acres']
    grand_total = sum(f['total_cost'] for f in fields.values())

    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, height - 50, "FarmTrack Season Cost Report")
    p.setFont("Helvetica", 12)
    p.drawString(50, height - 70, "Manitoba Farm Input Cost Summary")

    y = height - 120
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, y, f"Total Season Cost: ${grand_total:.2f}")
    y -= 40

    for field_name, field in fields.items():
        p.setFont("Helvetica-Bold", 13)
        p.drawString(50, y, f"Field: {field_name}")
        y -= 20
        p.setFont("Helvetica", 12)
        crop_line = f"Acres: {field['acres']}"
        if field.get('crop_type'):
            crop_line += f"  |  Crop: {field['crop_type']}"
        p.drawString(50, y, crop_line)
        y -= 20
        p.drawString(50, y, f"Total Cost: ${field['total_cost']:.2f}")
        y -= 20
        p.drawString(50, y, f"Cost Per Acre: ${field['cost_per_acre']:.2f}")
        y -= 30

    p.save()
    buffer.seek(0)

    return Response(buffer, mimetype='application/pdf',
        headers={'Content-Disposition': 'attachment;filename=farmtrack_report.pdf'})

@app.route('/equipment/export', methods=['GET', 'POST'])
@login_required
def export_equipment():
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    equip_file = f'{DATA_DIR}/equipment_{session["username"]}.json'
    maint_file = f'{DATA_DIR}/maintenance_{session["username"]}.json'

    if os.path.exists(equip_file):
        with open(equip_file, 'r') as f:
            equip_list = json.load(f)
    else:
        equip_list = []
    if os.path.exists(maint_file):
        with open(maint_file, 'r') as f:
            all_maint = json.load(f)
    else:
        all_maint = []

    if request.method == 'GET':
        return render_template('export_options.html', equip_list=equip_list)

    selected_columns = request.form.getlist('columns')
    if not selected_columns:
        selected_columns = ['date', 'service_type', 'hours', 'notes']

    column_labels = {'date': 'Date', 'service_type': 'Service Type', 'hours': 'Hours', 'notes': 'Notes'}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color='2C5F2E', end_color='2C5F2E', fill_type='solid')
    header_font_white = Font(name='Arial', bold=True, size=10, color='FFFFFF')

    if not equip_list:
        ws = wb.create_sheet('No Equipment')
        ws['A1'] = 'No equipment logged yet in FarmTrack.'
    else:
        for equip in equip_list:
            sheet_name = equip['name'][:31] if equip['name'] else f"Equipment {equip['id']}"
            ws = wb.create_sheet(sheet_name)

            ws['A1'] = 'Model'
            ws['B1'] = equip.get('model', '')
            ws['A2'] = 'Type'
            ws['B2'] = equip.get('equipment_type', '')
            ws['A3'] = 'Year'
            ws['B3'] = equip.get('year', '')
            ws['A4'] = 'Serial #'
            ws['B4'] = equip.get('serial_number', '')
            for r in range(1, 5):
                ws.cell(row=r, column=1).font = Font(name='Arial', bold=True, size=10)

            for col, key in enumerate(selected_columns, start=1):
                cell = ws.cell(row=6, column=col, value=column_labels.get(key, key))
                cell.font = header_font_white
                cell.fill = header_fill

            records = [m for m in all_maint if m['equipment_id'] == equip['id']]
            records.sort(key=lambda x: x.get('date', ''))

            row_num = 7
            for m in records:
                for col, key in enumerate(selected_columns, start=1):
                    ws.cell(row=row_num, column=col, value=m.get(key, ''))
                row_num += 1

            widths = {'date': 14, 'service_type': 22, 'hours': 10, 'notes': 30}
            from openpyxl.utils import get_column_letter
            for col, key in enumerate(selected_columns, start=1):
                ws.column_dimensions[get_column_letter(col)].width = widths.get(key, 16)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment;filename=farmtrack_equipment_maintenance.xlsx'})

if __name__ == '__main__':
    app.run(debug=True)
